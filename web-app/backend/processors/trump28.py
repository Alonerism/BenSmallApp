# Trump28 - Weekly to Cash & Payroll Processor
# Adapted from original Trump28.py to work with file bytes for web uploads
# Preserves EXACT cell reading/writing logic for client's hyper-specific Excel formats

import os
import re
from collections import defaultdict
from datetime import datetime
from io import BytesIO
from dataclasses import dataclass
from typing import Optional, Dict, List, Any

import pandas as pd
from openpyxl import load_workbook
from fuzzywuzzy import fuzz


def ws_to_df(ws):
    """Convert an openpyxl worksheet to a pandas DataFrame (for previews)."""
    rows = list(ws.values)
    if not rows:
        return pd.DataFrame()
    headers = [str(h) if h is not None else f"Col{i+1}" for i, h in enumerate(rows[0])]
    return pd.DataFrame(rows[1:], columns=headers)


def excel_first_sheet_to_df_bytes(file_bytes: bytes):
    """Read the first sheet of an Excel file (from bytes) into a DataFrame."""
    try:
        x = pd.ExcelFile(BytesIO(file_bytes), engine="openpyxl")
        sheet = x.sheet_names[0]
        return x.parse(sheet)
    except Exception:
        return pd.DataFrame()


@dataclass
class Trump28Result:
    cash_output_bytes: Optional[bytes]
    payroll_output_bytes: Optional[bytes]
    loans_output_bytes: Optional[bytes]
    counts: Dict[str, int]
    bonus_summary: Dict[str, Any]
    loan_notes: List[str]
    unmatched_reports: List[Dict[str, Any]]


def run_pipeline(
    weekly_bytes: bytes,
    cash_bytes: bytes,
    payroll_bytes: bytes,
    reimb_bytes: bytes,
    loans_bytes: Optional[bytes] = None,
    save: bool = True
) -> Trump28Result:
    """
    Core pipeline: reads Weekly, fills Cash & Payroll, computes totals/bonuses,
    applies reimbursements & loans, and returns outputs + a report dict.

    Accepts bytes instead of file paths for web compatibility.

    Changes from prior:
    - Weekly schema shifted left (old "Bonus Position" removed).
    - Bonus Position now from Reimbursements!C; Foreman uploads from Reimbursements!D (scales foreman bonus).
    - Category "c": SICK consumes the 24-hour Payroll cap first.
    - Loans.xlsx (optional) → deduct weekly payments from Cash Total (col F),
      update Open Loans (Total Paid/Balance), and move paid-off loans to History.
    """
    FLOOR_CASH_AT_ZERO = True

    # Load workbooks from bytes
    wb_reg = load_workbook(BytesIO(cash_bytes))
    ws_reg = wb_reg.active
    wb_ot = load_workbook(BytesIO(payroll_bytes))
    ws_ot = wb_ot.active
    wb_reimb = load_workbook(BytesIO(reimb_bytes), data_only=False)
    ws_reimb = wb_reimb.active

    # Weekly
    wb_weekly = pd.ExcelFile(BytesIO(weekly_bytes), engine="openpyxl")
    weekly_sheet = wb_weekly.sheet_names[0]
    df_weekly = wb_weekly.parse(weekly_sheet, header=None)

    if df_weekly.shape[1] < 17:
        raise ValueError(
            f"Weekly sheet '{weekly_sheet}' has only {df_weekly.shape[1]} columns; expected ≥ 17."
        )

    df = df_weekly.iloc[4:, :17].copy()
    df.columns = [
        'Name', 'Category',
        'Thu_Reg', 'Thu_OT', 'Fri_Reg', 'Fri_OT', 'Sat_Reg', 'Sat_OT',
        'Mon_Reg', 'Mon_OT', 'Tue_Reg', 'Tue_OT', 'Wed_Reg', 'Wed_OT',
        'Total_Reg', 'Total_OT', 'Sick'
    ]

    df['Name'] = df['Name'].astype(str).str.strip()
    for col in df.columns[2:14]:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    df['Total_OT']  = df[['Thu_OT','Fri_OT','Sat_OT','Mon_OT','Tue_OT','Wed_OT']].sum(axis=1)
    df['Total_Reg'] = df[['Thu_Reg','Fri_Reg','Sat_Reg','Mon_Reg','Tue_Reg','Wed_Reg']].sum(axis=1)

    # Matching helpers
    def last_name(name):
        toks = str(name).split()
        return toks[-1].lower() if toks else ""

    def first_token(name):
        toks = str(name).split()
        return toks[0].lower() if toks else ""

    def best_match(needle, haystack, min_score=92):
        wn = str(needle).strip()
        if not wn:
            return None, 0
        best, best_score = None, -1
        wn_last, wn_first = last_name(wn), first_token(wn)
        for c in haystack:
            score = fuzz.token_set_ratio(wn.lower(), c.lower())
            if score > best_score:
                best, best_score = c, score
            elif score == best_score and best is not None:
                c_last, b_last = last_name(c), last_name(best)
                c_first, b_first = first_token(c), first_token(best)
                if (c_last == wn_last) and (b_last != wn_last):
                    best = c
                elif (c_last == wn_last) and (b_last == wn_last):
                    if c_first[:1] == wn_first[:1] and b_first[:1] != wn_first[:1]:
                        best = c
        if best_score >= min_score:
            return best, best_score
        # fallback: same last name, lower bar
        fallback, fallback_score = None, -1
        for c in haystack:
            if last_name(c) == wn_last:
                score = fuzz.token_set_ratio(wn.lower(), c.lower())
                if score > fallback_score:
                    fallback, fallback_score = c, score
        if fallback and fallback_score >= 85:
            return fallback, fallback_score
        return None, 0

    cash_names = [str(ws_reg.cell(i, 1).value).strip()
                  for i in range(2, ws_reg.max_row + 1)
                  if ws_reg.cell(i, 1).value and str(ws_reg.cell(i, 1).value).strip()]
    payroll_names = [str(ws_ot.cell(i, 1).value).strip()
                     for i in range(2, ws_ot.max_row + 1)
                     if ws_ot.cell(i, 1).value and str(ws_ot.cell(i, 1).value).strip()]

    wk_to_cash = {}
    for nm in df['Name'].dropna().unique():
        c, _ = best_match(nm, cash_names, min_score=92)
        if c:
            wk_to_cash[nm] = c
    def cash_canon(nm): return wk_to_cash.get(nm)

    cash_to_payroll = {}
    for cn in set(wk_to_cash.values()):
        p, _ = best_match(cn, payroll_names, min_score=92)
        if p:
            cash_to_payroll[cn] = p
    def payroll_from_cash(cash_name): return cash_to_payroll.get(cash_name, cash_name)

    # Build unmatched reports
    skip_markers = {
        "employee name:", "nan", "* red coded absent", "reminders:",
        "payroll employees", "cash employees", "50/50 employees"
    }
    unmatched_reports = []
    for nm in df['Name']:
        if not nm or str(nm).strip().lower() in skip_markers:
            continue
        missing = []
        cash_match, _ = best_match(nm, cash_names, min_score=92)
        if not cash_match:
            missing.append("Cash")
        else:
            payroll_match, _ = best_match(cash_match, payroll_names, min_score=92)
            if not payroll_match:
                missing.append("Payroll")
        if missing:
            unmatched_reports.append({"name": str(nm), "missing": missing})

    # === Bonuses & reimbursements from Reimb sheet ===
    reg_yards = float(ws_reimb["B2"].value or 0)
    delfern_yards = float(ws_reimb["B3"].value or 0)
    total_yards = reg_yards + delfern_yards

    start_row = next(i for i in range(1, ws_reimb.max_row + 1)
                     if str(ws_reimb.cell(i,1).value).strip().lower() == "name") + 1

    def parse_uploads(val):
        if val is None:
            return 0.0
        if isinstance(val, (int, float)):
            return max(0.0, min(10.0, float(val)))
        s = str(val).strip()
        if not s:
            return 0.0
        parts = re.split(r"[^\d\.]+", s.replace(",", "+").replace(" ", ""))
        nums = []
        for p in parts:
            if p == "":
                continue
            try:
                nums.append(float(p))
            except:
                pass
        total = sum(nums) if nums else 0.0
        return max(0.0, min(10.0, total))

    # First pass: count foremen
    num_foremen = 0
    for i in range(start_row, ws_reimb.max_row + 1):
        role = ws_reimb.cell(i, 3).value
        if isinstance(role, str) and "foreman" in role.strip().lower():
            num_foremen += 1

    # Second pass: compute bonuses
    bonus_by_cash = defaultdict(float)
    people_with_bonus = 0
    for i in range(start_row, ws_reimb.max_row + 1):
        nm = ws_reimb.cell(i, 1).value
        role = ws_reimb.cell(i, 3).value
        uploads = ws_reimb.cell(i, 4).value
        if not nm or not isinstance(role, str) or not role.strip():
            continue

        role_l = role.strip().lower()
        bonus = 0.0
        if "foreman" in role_l and num_foremen > 0:
            bonus = total_yards / num_foremen
        elif "3x" in role_l:
            bonus = delfern_yards * 3
        elif "0.5" in role_l:
            bonus = total_yards * 0.5
        elif "1x" in role_l:
            bonus = total_yards * 1

        if bonus > 0:
            if "foreman" in role_l:
                uploads_total = parse_uploads(uploads)
                bonus *= (uploads_total / 10.0)
            cn, _ = best_match(str(nm).strip(), cash_names, min_score=90)
            if cn:
                bonus_by_cash[cn] += float(bonus)
                people_with_bonus += 1

    # Reimbursements (B)
    reimb_by_cash = defaultdict(float)
    for i in range(start_row, ws_reimb.max_row + 1):
        nm = ws_reimb.cell(i, 1).value
        raw = ws_reimb.cell(i, 2).value
        if not nm or str(nm).strip().lower() == "total":
            continue
        try:
            val = eval(raw[1:]) if isinstance(raw, str) and raw.startswith("=") else float(raw)
        except Exception:
            val = None
        if val is None:
            continue
        cn, _ = best_match(str(nm).strip(), cash_names, min_score=90)
        if cn:
            reimb_by_cash[cn] += float(val)

    def find_cash_row(sheet, name_str, type_val):
        for i in range(2, sheet.max_row + 1):
            n, t = sheet.cell(i, 1).value, sheet.cell(i, 2).value
            if n and t and str(n).strip() == str(name_str).strip() and str(t).strip().upper() == type_val.upper():
                return i
        return None

    def find_payroll_row(sheet, name_str, type_val):
        for i in range(2, sheet.max_row + 1):
            n, t = sheet.cell(i, 1).value, sheet.cell(i, 3).value
            if n and t and str(n).strip() == str(name_str).strip() and str(t).strip().upper() == type_val.upper():
                return i
        return None

    # === Fill hours ===
    sick_buffer = []
    for _, row in df.iterrows():
        wk_name = row["Name"]
        category = str(row["Category"]).strip().lower()
        reg = float(row["Total_Reg"])
        ot  = float(row["Total_OT"])
        if str(wk_name).lower() in skip_markers:
            continue
        cash_name = wk_to_cash.get(wk_name)
        if not cash_name:
            continue
        payroll_name = payroll_from_cash(cash_name)

        # Sick detection
        sick_hours = 0
        raw_row = df_weekly.iloc[row.name]
        for col_index in [2, 4, 6, 8, 10, 12]:
            if col_index < len(raw_row):
                raw_val = str(raw_row[col_index]).strip().lower()
                if 'sick' in raw_val:
                    sick_hours += 8
        if sick_hours > 0:
            sick_buffer.append((payroll_name, sick_hours))

        reg_row_cash = find_cash_row(ws_reg, cash_name, "R")
        ot_row_cash  = find_cash_row(ws_reg, cash_name, "OT")
        reg_row_pay  = find_payroll_row(ws_ot, payroll_name, "R")

        if category == "a":  # Full Payroll
            if reg_row_pay:
                ws_ot.cell(reg_row_pay, 4).value = reg
            if ot_row_cash:
                ws_reg.cell(ot_row_cash, 3).value = ot

        elif category == "b":  # All Cash
            reg_capped = min(reg, 40)
            reg_overflow = max(0, reg - 40)
            total_ot = reg_overflow + ot
            if reg_row_cash:
                ws_reg.cell(reg_row_cash, 3).value = reg_capped
            if ot_row_cash:
                ws_reg.cell(ot_row_cash, 3).value = total_ot

        elif category == "c":  # Split: Payroll cap 24, sick consumes cap first
            CAP = 24
            cap_after_sick = max(0, CAP - sick_hours)
            payroll_r = min(reg, cap_after_sick)

            cash_reg = max(0, reg - payroll_r)
            reg_capped = min(cash_reg, 40)
            reg_overflow = max(0, cash_reg - 40)
            total_ot = reg_overflow + ot

            if reg_row_pay:
                ws_ot.cell(reg_row_pay, 4).value = payroll_r
            if reg_row_cash:
                ws_reg.cell(reg_row_cash, 3).value = reg_capped
            if ot_row_cash:
                ws_reg.cell(ot_row_cash, 3).value = total_ot

    # Write sick hours
    for pname, sh in sick_buffer:
        sick_row = find_payroll_row(ws_ot, pname, "SICK")
        if sick_row:
            ws_ot.cell(sick_row, 4).value = sh

    # === Row pay & totals on Cash ===
    already_seen_for_E = set()
    pay_totals = defaultdict(lambda: {'reg': 0.0, 'ot': 0.0})

    for i in range(2, ws_reg.max_row + 1):
        name = ws_reg.cell(i, 1).value
        typ  = ws_reg.cell(i, 2).value
        hrs  = ws_reg.cell(i, 3).value
        rate = ws_reg.cell(i, 4).value
        if not name:
            continue

        typ_norm = (str(typ).strip().upper() if typ is not None else "")

        try:
            h = float(hrs)
        except Exception:
            h = 0.0
        try:
            r = float(str(rate).replace("$", "").replace(",", "").strip())
        except Exception:
            r = 0.0

        pay = round(h * r, 2)

        if name not in already_seen_for_E:
            ws_reg.cell(i, 5).value = pay
            already_seen_for_E.add(name)
        else:
            ws_reg.cell(i, 5).value = None

        if typ_norm == "R":
            pay_totals[name]['reg'] += pay
        elif typ_norm == "OT":
            pay_totals[name]['ot'] += pay

    # === Pre-loan totals per person ===
    preloan_total_by_name = {}
    names_seen_in_cash = set()
    for i in range(2, ws_reg.max_row + 1):
        nm = ws_reg.cell(i, 1).value
        if not nm or nm in names_seen_in_cash:
            continue
        base_total = pay_totals[nm]['reg'] + pay_totals[nm]['ot']
        bonus = bonus_by_cash.get(nm, 0.0)
        reimb = reimb_by_cash.get(nm, 0.0)
        preloan_total_by_name[nm] = round(base_total + bonus + reimb, 2)
        names_seen_in_cash.add(nm)

    # === LOANS ===
    loan_deducted_by_cash = defaultdict(float)
    loan_notes = []
    loans_summary = {"processed": 0, "closed": 0}
    wb_loans = None

    def _parse_money(val):
        if val is None or (isinstance(val, str) and not val.strip()):
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).strip()
        try:
            if s.startswith("="):
                return float(eval(s[1:]))
            return float(s.replace(",", ""))
        except Exception:
            return 0.0

    if loans_bytes:
        wb_loans = load_workbook(BytesIO(loans_bytes))
        ws_open = wb_loans.worksheets[0]
        ws_hist = wb_loans.worksheets[1] if len(wb_loans.worksheets) > 1 else wb_loans.create_sheet("HISTORY")

        open_headers = {str(ws_open.cell(1, c).value).strip().lower(): c
                        for c in range(1, ws_open.max_column + 1)
                        if ws_open.cell(1, c).value}

        def _ocol(search, default=None):
            for k, c in open_headers.items():
                if search in k:
                    return c
            return default

        name_col       = _ocol("name", 1)
        payment_col    = _ocol("payment", 3)
        amount_col     = _ocol("loan amount", 2)
        date_taken_col = _ocol("date", None)
        total_paid_col = _ocol("total paid", None)
        balance_col    = _ocol("balance", None)

        per_person_rows = defaultdict(list)
        for r in range(2, ws_open.max_row + 1):
            nm_cell = ws_open.cell(r, name_col).value if name_col else None
            if not nm_cell or str(nm_cell).strip() == "":
                continue
            intended = _parse_money(ws_open.cell(r, payment_col).value) if payment_col else 0.0
            if intended <= 0:
                continue

            nm_str = str(nm_cell).strip()
            cn, _ = best_match(nm_str, cash_names, min_score=90)
            if not cn:
                continue

            loan_amt  = _parse_money(ws_open.cell(r, amount_col).value) if amount_col else 0.0
            prev_paid = _parse_money(ws_open.cell(r, total_paid_col).value) if total_paid_col else 0.0

            if balance_col:
                bal_cell = ws_open.cell(r, balance_col).value
                try:
                    start_bal = float(bal_cell) if bal_cell not in (None, "") else (loan_amt - prev_paid)
                except Exception:
                    start_bal = loan_amt - prev_paid
            else:
                start_bal = loan_amt - prev_paid

            per_person_rows[cn].append({
                "r": r,
                "display_name": nm_str,
                "intended": round(intended, 2),
                "start_bal": round(start_bal, 2),
                "loan_amt": round(loan_amt, 2),
                "prev_paid": round(prev_paid, 2),
                "date_taken": (ws_open.cell(r, date_taken_col).value if date_taken_col else None),
            })
            loans_summary["processed"] += 1

        for person, rows in per_person_rows.items():
            available = max(0.0, preloan_total_by_name.get(person, 0.0))
            total_intended = sum(x["intended"] for x in rows)
            actually_taken_total = 0.0

            for info in rows:
                r          = info["r"]
                intended   = info["intended"]
                start_bal  = max(0.0, info["start_bal"])

                pay_bal_capped = min(intended, start_bal)
                take_now = min(pay_bal_capped, max(0.0, available - actually_taken_total))

                if take_now > 0:
                    new_paid = info["prev_paid"] + take_now
                    new_bal  = max(0.0, start_bal - take_now)

                    if total_paid_col:
                        ws_open.cell(r, total_paid_col).value = round(new_paid, 2)
                    if balance_col:
                        ws_open.cell(r, balance_col).value = round(new_bal, 2)

                    actually_taken_total += take_now

                    if new_bal <= 0.000001:
                        loans_summary["closed"] += 1
                        hist_headers = {str(ws_hist.cell(1, c).value).strip().lower(): c
                                        for c in range(1, ws_hist.max_column + 1)
                                        if ws_hist.cell(1, c).value}
                        def _hcol(search, default):
                            for k, c in hist_headers.items():
                                if search in k:
                                    return c
                            return default

                        h_name = _hcol("name", 1)
                        h_amt  = _hcol("loan amount", 2)
                        h_pay  = _hcol("payment", 3)
                        h_date = _hcol("date", 4)

                        insert_at = 3
                        ws_hist.insert_rows(insert_at)
                        ws_hist.cell(insert_at, h_name).value = info["display_name"]
                        ws_hist.cell(insert_at, h_amt).value  = info["loan_amt"]
                        ws_hist.cell(insert_at, h_pay).value  = take_now
                        ws_hist.cell(insert_at, h_date).value = info["date_taken"]

                        for c in range(1, ws_open.max_column + 1):
                            ws_open.cell(r, c).value = None

                if intended > start_bal + 1e-9:
                    loan_notes.append(f"{person}: capped to balance ${start_bal:.2f} (intended ${intended:.2f})")
                if take_now + 1e-9 < min(intended, start_bal):
                    short = min(intended, start_bal) - take_now
                    loan_notes.append(f"{person}: only ${take_now:.2f} deducted, ${short:.2f} rolled")

            loan_deducted_by_cash[person] = round(actually_taken_total, 2)

    # === Totals in Cash (F) with bonuses, reimb, and loan deductions ===
    already_totaled = set()
    for i in range(2, ws_reg.max_row + 1):
        name = ws_reg.cell(i, 1).value
        if not name or name in already_totaled:
            continue

        base_total = pay_totals[name]['reg'] + pay_totals[name]['ot']
        bonus = bonus_by_cash.get(name, 0.0)
        reimb = reimb_by_cash.get(name, 0.0)
        subtotal = base_total + bonus + reimb

        loan_deduction = loan_deducted_by_cash.get(name, 0.0)
        net_total = round(subtotal - loan_deduction, 2)
        if FLOOR_CASH_AT_ZERO:
            net_total = max(0.0, net_total)

        ws_reg.cell(i, 6).value = net_total
        already_totaled.add(name)

    # Save to bytes
    cash_output_bytes = None
    payroll_output_bytes = None
    loans_output_bytes = None

    if save:
        cash_buffer = BytesIO()
        wb_reg.save(cash_buffer)
        cash_output_bytes = cash_buffer.getvalue()

        payroll_buffer = BytesIO()
        wb_ot.save(payroll_buffer)
        payroll_output_bytes = payroll_buffer.getvalue()

        if wb_loans:
            loans_buffer = BytesIO()
            wb_loans.save(loans_buffer)
            loans_output_bytes = loans_buffer.getvalue()

    return Trump28Result(
        cash_output_bytes=cash_output_bytes,
        payroll_output_bytes=payroll_output_bytes,
        loans_output_bytes=loans_output_bytes,
        counts={
            "weekly_rows": int(df.shape[0]),
            "sick_entries": len(sick_buffer),
            "loans_processed": loans_summary["processed"] if loans_bytes else 0,
            "loans_closed": loans_summary["closed"] if loans_bytes else 0,
        },
        bonus_summary={
            "reg_yards": reg_yards,
            "delfern_yards": delfern_yards,
            "total_yards": total_yards,
            "num_foremen": num_foremen,
        },
        loan_notes=loan_notes,
        unmatched_reports=unmatched_reports,
    )
