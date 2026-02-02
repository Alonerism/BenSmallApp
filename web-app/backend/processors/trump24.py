# Trump24 - Full Week Processor
# Adapted from original Trump24.py (time_to_weekly.py) to work with file bytes for web uploads
# Preserves EXACT cell reading/writing logic for client's hyper-specific Excel formats

import os, re, math, datetime
from collections import defaultdict, Counter
from dataclasses import dataclass
from typing import Dict, Tuple, Optional
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook
from fuzzywuzzy import fuzz
from functools import lru_cache


# =======================
# Precompiled regexes
# =======================
TIME_HHMM_RE = re.compile(r'^(\d+):(\d{2})(?::\d{2})?$')
DATE_HEADER_RE = re.compile(r"Timecard Date:\s*(\d{1,2}/\d{1,2}/\d{4})")

# =======================
# Helpers
# =======================
def _parse_hours_cell(x):
    """Accepts numbers, 'H:MM' or 'HH:MM(:SS)' or '8.0' → float hours."""
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return None if pd.isna(x) else float(x)
    s = str(x).strip()
    if not s:
        return None
    m = TIME_HHMM_RE.match(s)
    if m:
        h = int(m.group(1)); mm = int(m.group(2))
        return h + mm/60.0
    try:
        return float(s)
    except:
        return None

def _round_to(x, step=0.5):
    return round(x / step) * step

@lru_cache(maxsize=4096)
def _norm(s: str) -> str:
    return re.sub(r"[^a-z\s\-']", " ", str(s).lower()).strip()

def _best_match(needle, haystack, min_score, fallback_score):
    wn = _norm(needle)
    best, score = None, -1
    for c in haystack:
        sc = fuzz.token_set_ratio(wn, _norm(c))
        if sc > score:
            best, score = c, sc
    if score >= min_score:
        return best, score
    wn_last = wn.split()[-1] if wn.split() else ""
    fallback, fscore = None, -1
    for c in haystack:
        cn = _norm(c)
        c_last = cn.split()[-1] if cn.split() else ""
        if c_last == wn_last:
            sc = fuzz.token_set_ratio(wn, cn)
            if sc > fscore:
                fallback, fscore = c, sc
    if fallback and fscore >= fallback_score:
        return fallback, fscore
    return None, 0

def _is_weekday(d: datetime.date) -> bool:
    return d.weekday() < 5

def _fmt_date(v):
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%m/%d/%Y")
    try:
        dt = pd.to_datetime(v, errors="coerce")
        return "" if pd.isna(dt) else dt.strftime("%m/%d/%Y")
    except Exception:
        return str(v)

def _fmt_hhmm(hours):
    """Convert float hours → H:MM (e.g., 7.98 -> 7:59). NaN/None -> ''."""
    if hours is None:
        return ""
    try:
        h = float(hours)
    except Exception:
        return "" if str(hours).strip() == "" else str(hours)
    if math.isnan(h) or math.isinf(h):
        return ""
    neg = h < 0
    h = abs(h)
    H = int(math.floor(h))
    M = int(round((h - H) * 60))
    if M == 60:
        H += 1
        M = 0
    return f"{'-' if neg else ''}{H}:{M:02d}"

def _reason_keys(reason_str):
    """'rounded(8.05->8.00), low_name_match(0)' -> {'rounded','low_name_match'}"""
    keys = set()
    if isinstance(reason_str, str) and reason_str.strip():
        for part in [p.strip() for p in reason_str.split(",") if p.strip()]:
            key = re.sub(r"\(.*\)", "", part).strip()
            if key:
                keys.add(key)
    return keys

def _mins_from_hours(delta_hours):
    """Float hours → integer minutes (rounded)."""
    try:
        return int(round(float(delta_hours) * 60))
    except Exception:
        return None


def _build_longest_stint_views(
    *,
    rows_in_block: dict,
    hours_lookup: dict,
    tar_to_wk: Dict[str, Tuple[str,int]],
    long_stint_flag: float,
    round_to_hours: float,
    suggest_lunch_deduct: float,
):
    """
    Returns:
      by_day_df: per day/person longest stint (with flags/suggestions)
      leader_df: per person weekly leaderboard
    """
    recs = []
    for (d, emp), stints in rows_in_block.items():
        if not stints:
            continue
        longest = max([s for s in stints if isinstance(s, (int,float)) and not math.isnan(s)] or [None])
        if longest is None:
            continue
        rounded_total = hours_lookup.get((d, emp), None)
        wmatch, _ = tar_to_wk.get(emp, (None, 0))
        flagged = bool(longest >= long_stint_flag)
        suggested = None
        if flagged and isinstance(rounded_total, (int, float)):
            suggested = max(_round_to(rounded_total - suggest_lunch_deduct, round_to_hours), 0.0)

        recs.append({
            "Date": d.strftime("%m/%d/%Y"),
            "TAR_Name": emp,
            "Weekly_Name": wmatch or "",
            "LongestStint": float(longest),
            "LongestStint_str": _fmt_hhmm(longest),
            "RoundedHours": rounded_total if rounded_total is not None else "",
            "RoundedHours_str": _fmt_hhmm(rounded_total) if rounded_total is not None else "",
            "Flag_LongStint": "Y" if flagged else "",
            "SuggestedHours": suggested if suggested is not None else "",
            "SuggestedHours_str": _fmt_hhmm(suggested) if suggested is not None else "",
        })

    by_day_df = pd.DataFrame(recs)
    if not by_day_df.empty:
        by_day_df["__flag"] = (by_day_df["Flag_LongStint"] == "Y").astype(int)
        by_day_df = by_day_df.sort_values(
            ["__flag","LongestStint","Weekly_Name","TAR_Name","Date"],
            ascending=[False, False, True, True, True]
        ).drop(columns="__flag")

        wk  = by_day_df["Weekly_Name"].astype(str)
        tar = by_day_df["TAR_Name"].astype(str)
        person_key = wk.where(wk.str.strip() != "", tar)

        grp = by_day_df.groupby(person_key, dropna=False)
        leader = grp.apply(lambda g: pd.Series({
            "Person": (g["Weekly_Name"].iloc[0] or g["TAR_Name"].iloc[0]),
            "MaxLongestStint": g["LongestStint"].max(),
            "MaxLongestStint_str": _fmt_hhmm(g["LongestStint"].max()),
            "DateOfMax": g.loc[g["LongestStint"].idxmax(), "Date"],
            "Days_Over_Threshold": int((g["Flag_LongStint"] == "Y").sum()),
            "Week_Rounded_Total": round(float(pd.to_numeric(g["RoundedHours"], errors="coerce").sum() or 0.0), 2),
        })).reset_index(drop=True)

        leader = leader.sort_values(
            ["MaxLongestStint","Days_Over_Threshold","Person"],
            ascending=[False, False, True]
        )
    else:
        leader = pd.DataFrame(columns=[
            "Person","MaxLongestStint","MaxLongestStint_str","DateOfMax",
            "Days_Over_Threshold","Week_Rounded_Total"
        ])

    return by_day_df, leader


def _build_preview_compact_lists(
    *,
    rows_in_block: dict,
    hours_lookup: dict,
    tar_to_wk: Dict[str, Tuple[str,int]],
    round_to_hours: float,
    reg_cap: float,
    long_stint_flag: float,
    suggest_lunch_deduct: float,
    min_stint_for_list: float = 4.0,
    ot_threshold: float = 8.0
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Returns:
      longest_compact: one row per person (their longest single stint >= min_stint_for_list)
      ot_compact:      one row per person (their max daily total > ot_threshold)
    """
    recs = []
    for (d, emp), stints in rows_in_block.items():
        if not stints:
            continue
        longest = max([s for s in stints if isinstance(s, (int, float)) and not math.isnan(s)] or [None])
        total = hours_lookup.get((d, emp), None)

        wmatch, _ = tar_to_wk.get(emp, (None, 0))
        person = (wmatch or str(emp or "")).strip()
        if not person:
            continue
        low = person.lower()
        if low.startswith("total") or "grand total" in low:
            continue

        recs.append({
            "Date": d,
            "Person": person,
            "TAR_Name": str(emp),
            "Weekly_Name": wmatch or "",
            "LongestStint": float(longest) if longest is not None else None,
            "DayTotal": float(total) if isinstance(total, (int, float)) else None,
        })

    base = pd.DataFrame(recs)
    if base.empty:
        return (pd.DataFrame(columns=["Person","Date","LongestStint","LongestStint_str","DayTotal","DayTotal_str","SuggestedHours","SuggestedHours_str"]),
                pd.DataFrame(columns=["Person","Date","DayTotal","DayTotal_str","OT_Hours","OT_Hours_str","MaxSingleStint","MaxSingleStint_str"]))

    # A) Longest single stint per person
    a = base.dropna(subset=["LongestStint"]).copy()
    a = a[a["LongestStint"] >= float(min_stint_for_list)]
    if not a.empty:
        a = a.sort_values(["Person","LongestStint","DayTotal","Date"],
                          ascending=[True, False, False, True])
        a = a.drop_duplicates(subset=["Person"], keep="first")
        def _suggest(row):
            if row["LongestStint"] is not None and row["LongestStint"] >= long_stint_flag and isinstance(row["DayTotal"], (int,float)):
                return max(_round_to(row["DayTotal"] - suggest_lunch_deduct, round_to_hours), 0.0)
            return None
        a["SuggestedHours"] = a.apply(_suggest, axis=1)
        a["LongestStint_str"]  = a["LongestStint"].apply(_fmt_hhmm)
        a["DayTotal_str"]      = a["DayTotal"].apply(_fmt_hhmm)
        a["SuggestedHours_str"]= a["SuggestedHours"].apply(lambda x: _fmt_hhmm(x) if x is not None else "")
        a = a.sort_values(["LongestStint","DayTotal","Person"], ascending=[False, False, True])
        longest_compact = a[["Person","Date","LongestStint","LongestStint_str","DayTotal","DayTotal_str","SuggestedHours","SuggestedHours_str"]].copy()
        longest_compact["Date"] = longest_compact["Date"].apply(_fmt_date)
    else:
        longest_compact = pd.DataFrame(columns=["Person","Date","LongestStint","LongestStint_str","DayTotal","DayTotal_str","SuggestedHours","SuggestedHours_str"])

    # B) Max OT day per person
    b = base.dropna(subset=["DayTotal"]).copy()
    b = b[b["DayTotal"] > float(ot_threshold)]
    if not b.empty:
        b = b.sort_values(["Person","DayTotal","LongestStint","Date"],
                          ascending=[True, False, False, True])
        b = b.drop_duplicates(subset=["Person"], keep="first")
        b["OT_Hours"] = (b["DayTotal"] - float(reg_cap)).clip(lower=0.0)
        b["DayTotal_str"]     = b["DayTotal"].apply(_fmt_hhmm)
        b["OT_Hours_str"]     = b["OT_Hours"].apply(_fmt_hhmm)
        b["MaxSingleStint_str"]= b["LongestStint"].apply(_fmt_hhmm)
        b = b.sort_values(["DayTotal","Person"], ascending=[False, True])
        ot_compact = b[["Person","Date","DayTotal","DayTotal_str","OT_Hours","OT_Hours_str","LongestStint","MaxSingleStint_str"]].copy()
        ot_compact.rename(columns={"LongestStint":"MaxSingleStint"}, inplace=True)
        ot_compact["Date"] = ot_compact["Date"].apply(_fmt_date)
    else:
        ot_compact = pd.DataFrame(columns=["Person","Date","DayTotal","DayTotal_str","OT_Hours","OT_Hours_str","MaxSingleStint","MaxSingleStint_str"])

    return longest_compact, ot_compact


# =======================
# Core parsing + mapping
# =======================
def _parse_time_activity(raw_times_bytes: bytes):
    """
    Accepts bytes instead of file path.
    Returns:
        daily_long: DataFrame with columns [Date, Employee, RawHours, RoundedHours]
        rows_in_block: dict[(date, employee)] -> list of stints (floats)
    """
    xl = pd.ExcelFile(BytesIO(raw_times_bytes), engine="openpyxl")
    df = xl.parse(xl.sheet_names[0], header=None)

    current_date = None
    rows_in_block = defaultdict(list)
    records = []

    for _, row in df.iterrows():
        found = None
        for cell in row.tolist():
            if isinstance(cell, str):
                m = DATE_HEADER_RE.search(cell)
                if m:
                    found = m.group(1); break
        if found:
            current_date = datetime.datetime.strptime(found, "%m/%d/%Y").date()
            continue
        if current_date is None:
            continue
        if isinstance(row.iloc[0], str) and row.iloc[0].strip().lower() == "employee":
            continue

        emp = row.iloc[0] if row.shape[0] > 0 else None
        if not emp or str(emp).strip().lower() == "nan":
            continue
        hrs = _parse_hours_cell(row.iloc[5] if row.shape[0] > 5 else None)
        if hrs is None or pd.isna(hrs) or hrs <= 0:
            continue

        emp_str = str(emp).strip()
        low = emp_str.lower()
        if low.startswith("total") or "grand total" in low:
            continue

        records.append({"Date": current_date, "Employee": emp_str, "RawHours": float(hrs)})
        rows_in_block[(current_date, emp_str)].append(float(hrs))

    if not records:
        raise RuntimeError("Parsed zero rows from Time Activity Report. Check that column A=Employee and column F=Total Hours.")

    daily = pd.DataFrame(records).groupby(["Date","Employee"], as_index=False)["RawHours"].sum()
    return daily, rows_in_block


def _read_weekly_structure(weekly_template_bytes: bytes, *, open_wb: bool = True):
    """
    Accepts bytes instead of file path.
    Returns:
        day_map: {date -> {"reg_col": j, "ot_col": j+1, "header": str}}
        weekly_rows: list[(row_index0, name_str)]
        start_year: int
        wb, ws: openpyxl workbook & active sheet (or None if open_wb=False)
        wk_df: pandas' sheet (for potential debug/preview)
    """
    xl = pd.ExcelFile(BytesIO(weekly_template_bytes), engine="openpyxl")
    sheet = xl.sheet_names[0]
    wk_df = xl.parse(sheet, header=None)

    week_of_row = None
    for i, v in enumerate(wk_df.iloc[:,0]):
        if isinstance(v, str) and "week of" in v.lower():
            week_of_row = i; break
    if week_of_row is None:
        raise RuntimeError("Couldn't find a 'Week Of :' row in WeeklyTime.")

    week_str = str(wk_df.iat[week_of_row, 0])
    m = re.search(r'(\d{2})\.(\d{2})\.(\d{2})\s*-\s*(\d{2})\.(\d{2})\.(\d{2})', week_str)
    start_year = 2000 + int(m.group(3)) if m else datetime.date.today().year

    day_hdr_row = week_of_row + 1
    sub_hdr_row = day_hdr_row + 1

    day_map = {}
    for j in range(wk_df.shape[1]):
        val = wk_df.iat[day_hdr_row, j]
        if isinstance(val, str) and re.search(r'\d{2}/\d{2}', val):
            mm, dd = map(int, re.search(r'(\d{2})/(\d{2})', val).groups())
            date = datetime.date(start_year, mm, dd)
            reg_ok = isinstance(wk_df.iat[sub_hdr_row, j], str) and 'reg' in wk_df.iat[sub_hdr_row, j].lower()
            ot_ok  = isinstance(wk_df.iat[sub_hdr_row, j+1], str) and 'ot'  in wk_df.iat[sub_hdr_row, j+1].lower()
            if reg_ok and ot_ok:
                day_map[date] = {"reg_col": j, "ot_col": j+1, "header": val}

    if not day_map:
        raise RuntimeError("Couldn't map any day columns in WeeklyTime.")

    start_row = None
    for i, v in enumerate(wk_df.iloc[:,0]):
        if isinstance(v, str) and v.strip().lower().startswith("employee name"):
            start_row = i + 1; break
    if start_row is None:
        raise RuntimeError("Couldn't find 'Employee Name:' header row in WeeklyTime.")

    weekly_rows = []
    for r in range(start_row, wk_df.shape[0]):
        name_cell = wk_df.iat[r, 0]
        if name_cell is None or (isinstance(name_cell, float) and pd.isna(name_cell)):
            continue
        s = str(name_cell).strip()
        if s and s.lower() != "nan":
            weekly_rows.append((r, s))

    if open_wb:
        wb = load_workbook(BytesIO(weekly_template_bytes))
        ws = wb[wb.sheetnames[0]]
    else:
        wb = None
        ws = None
    return day_map, weekly_rows, start_year, wb, ws, wk_df


# =======================
# Result dataclass
# =======================
@dataclass
class TimeToWeeklyResult:
    output_bytes: Optional[bytes]
    filled_cells: int
    review_df: pd.DataFrame
    name_matching: pd.DataFrame
    daily_long: pd.DataFrame
    secretary_message: str
    review_df_display: Optional[pd.DataFrame] = None
    daily_long_display: Optional[pd.DataFrame] = None
    violations_by_person: Optional[pd.DataFrame] = None
    longest_stints_by_day: Optional[pd.DataFrame] = None
    longest_stints_leaderboard: Optional[pd.DataFrame] = None
    preview_longest_shifts: Optional[pd.DataFrame] = None
    preview_ot_days: Optional[pd.DataFrame] = None


def run_time_to_weekly(
    raw_times_bytes: bytes,
    weekly_template_bytes: bytes,
    *,
    round_to_hours: float       = 0.5,
    reg_cap: float              = 8.0,
    daily_max_hours: float      = 16.0,
    long_stint_flag: float      = 10.0,
    match_min_score: int        = 92,
    fallback_score: int         = 85,
    flag_low_weekday: float     = 2.0,
    suggest_lunch_deduct: float = 0.5,
    save: bool = True,
) -> TimeToWeeklyResult:
    """
    Process full week Time Activity Report and fill entire WeeklyTime template.
    Accepts bytes instead of file paths for web compatibility.
    """
    # 1) Parse raw times
    daily, rows_in_block = _parse_time_activity(raw_times_bytes)
    daily = daily.copy()
    daily["RoundedHours"] = daily["RawHours"].apply(lambda x: _round_to(x, round_to_hours))

    # 2) Read Weekly structure
    day_map, weekly_rows, start_year, wb, ws, wk_df = _read_weekly_structure(
        weekly_template_bytes, open_wb=save
    )

    # 3) Name matching
    weekly_names = [n for _, n in weekly_rows]
    tar_names = sorted(daily["Employee"].astype(str).unique())

    match_rows = []
    wk_to_tar: Dict[str, Tuple[str, int]] = {}
    tar_to_wk: Dict[str, Tuple[str, int]] = {}

    for tn in tar_names:
        wmatch, score = _best_match(tn, weekly_names, match_min_score, fallback_score)
        match_rows.append({
            "TAR Name": tn,
            "Weekly Match": wmatch or "",
            "Score": score,
            "Flag": "" if score >= match_min_score else "REVIEW"
        })
        if wmatch:
            tar_to_wk[tn] = (wmatch, score)

    for tn, (wmatch, score) in tar_to_wk.items():
        if wmatch not in wk_to_tar or score > wk_to_tar[wmatch][1]:
            wk_to_tar[wmatch] = (tn, score)

    name_matching = pd.DataFrame(match_rows).sort_values(
        ["Flag","Score","Weekly Match","TAR Name"], ascending=[True, False, True, True]
    )

    # 4) Build review queue + lookups
    review = []
    hours_lookup = {}
    segments_lookup = {}
    for _, row in daily.iterrows():
        d, tn = row["Date"], row["Employee"]
        rounded = float(row["RoundedHours"])
        hours_lookup[(d, tn)] = rounded
        segments_lookup[(d, tn)] = rows_in_block.get((d, tn), [])

    days_worked = Counter()
    for (d, tn), h in hours_lookup.items():
        if h > 0:
            days_worked[tn] += 1
    worked_regulars = {tn for tn, cnt in days_worked.items() if cnt >= 2}

    def add_review_row(d, tn, wmatch, score, raw_val, rounded_val, stints, reasons, suggested):
        review.append({
            "Date": d.strftime("%m/%d/%Y"),
            "TAR_Name": tn,
            "Weekly_Name": wmatch or "",
            "MatchScore": score,
            "Segments": ";".join(f"{s:.2f}" for s in stints) if stints else "",
            "RawHours": round(raw_val, 2) if raw_val is not None else "",
            "RoundedHours": round(rounded_val, 2) if rounded_val is not None else "",
            "SuggestedHours": suggested if suggested is not None else "",
            "Reasons": ", ".join(reasons) if reasons else ""
        })

    for (d, tn), rounded in hours_lookup.items():
        wmatch, score = tar_to_wk.get(tn, (None, 0))
        reasons, suggested = [], None
        raw_series = daily[(daily["Date"]==d)&(daily["Employee"]==tn)]["RawHours"]
        raw = float(raw_series.iloc[0]) if not raw_series.empty else None

        if score < match_min_score:
            reasons.append(f"low_name_match({score})")
        if rounded is not None and rounded > daily_max_hours:
            reasons.append(f"gt_daily_max({rounded})")
        if rounded is not None and (0 < rounded <= flag_low_weekday) and _is_weekday(d):
            reasons.append(f"very_low_weekday({rounded})")
        if raw is not None and rounded is not None and abs(raw - rounded) >= 0.01:
            reasons.append(f"rounded({raw:.2f}->{rounded:.2f})")

        stints = segments_lookup.get((d, tn), [])
        if len(stints) == 1 and stints[0] >= long_stint_flag:
            reasons.append(f"single_long_stint({stints[0]:.2f}h)")
            suggested = max(_round_to(rounded - suggest_lunch_deduct, round_to_hours), 0.0)

        if reasons or suggested is not None:
            add_review_row(d, tn, wmatch, score, raw, rounded, stints, reasons, suggested)

    mapped_dates = set(day_map.keys())
    for _, wname in weekly_rows:
        tn, score = wk_to_tar.get(wname, (None, 0))
        if not tn or score < match_min_score or tn not in worked_regulars:
            continue
        for d in mapped_dates:
            if (d, tn) not in hours_lookup:
                add_review_row(d, tn, wname, score, None, 0.0, [], ["missing_day_for_regular"], None)

    review_df = pd.DataFrame(review).sort_values(["Weekly_Name","Date"]).reset_index(drop=True)

    # ---------- Pretty preview frames (HH:MM) & violation ordering ----------
    rev_disp = review_df.copy()
    violations_by_person = pd.DataFrame()

    if not rev_disp.empty:
        def _longest_stint(seg_str):
            tok = str(seg_str or "").strip()
            vals = []
            if tok:
                for t in tok.split(";"):
                    t = t.strip()
                    if not t:
                        continue
                    try:
                        v = float(t)
                        if not math.isnan(v):
                            vals.append(v)
                    except:
                        pass
            return max(vals) if vals else None

        rev_disp["RawHours_str"]       = rev_disp["RawHours"].apply(lambda x: _fmt_hhmm(x) if pd.notna(x) and x != "" else "")
        rev_disp["RoundedHours_str"]   = rev_disp["RoundedHours"].apply(lambda x: _fmt_hhmm(x) if pd.notna(x) and x != "" else "")
        rev_disp["SuggestedHours_str"] = rev_disp["SuggestedHours"].apply(lambda x: _fmt_hhmm(x) if pd.notna(x) and x != "" else "")
        rev_disp["LongestStint"]       = rev_disp["Segments"].apply(_longest_stint)
        rev_disp["LongestStint_str"]   = rev_disp["LongestStint"].apply(lambda x: _fmt_hhmm(x) if pd.notna(x) else "")
        rev_disp["Person"]             = rev_disp.apply(lambda r: (r["Weekly_Name"] or r["TAR_Name"]), axis=1)

        def _effective_reason_keys(row):
            keys = _reason_keys(row.get("Reasons", ""))
            rounded_val = row.get("RoundedHours", None)
            if isinstance(rounded_val, (int, float)) and abs(float(rounded_val) - 8.0) < 1e-9:
                keys = {k for k in keys if k != "rounded"}
            return keys

        rev_disp["ReasonKeys"] = rev_disp.apply(_effective_reason_keys, axis=1)

        WEIGHT = {
            "low_name_match": 2,
            "single_long_stint": 3,
            "gt_daily_max": 3,
            "very_low_weekday": 1,
            "missing_day_for_regular": 1,
            "rounded": 0,
        }
        def _row_weight(keys):
            return sum(WEIGHT.get(k, 1) for k in keys)

        rev_disp["ViolationWeight"] = rev_disp["ReasonKeys"].apply(_row_weight)

        by_person = []
        for person, grp in rev_disp.groupby("Person", dropna=False):
            total_score = int(grp["ViolationWeight"].sum())
            rows_count  = int((grp["ViolationWeight"] > 0).sum())
            rc = Counter()
            for ks in grp["ReasonKeys"]:
                rc.update(ks)
            by_person.append({
                "Person": person,
                "ViolationScore": total_score,
                "ViolationRows": rows_count,
                "TopReasons": ", ".join(f"{k}:{v}" for k, v in rc.most_common(5))
            })
        violations_by_person = pd.DataFrame(by_person).sort_values(
            ["ViolationScore","ViolationRows","Person"], ascending=[False, False, True]
        )

        score_map = violations_by_person.set_index("Person")["ViolationScore"].to_dict()
        rows_map  = violations_by_person.set_index("Person")["ViolationRows"].to_dict()
        rev_disp["PersonScore"] = rev_disp["Person"].map(score_map).fillna(0).astype(int)
        rev_disp["PersonRows"]  = rev_disp["Person"].map(rows_map).fillna(0).astype(int)

        def _pretty_reasons(r):
            s = str(r.get("Reasons","") or "")
            s = re.sub(r"rounded\(([-\d\.]+)->([-\d\.]+)\)", lambda m: f"rounded({_fmt_hhmm(m.group(1))}->{_fmt_hhmm(m.group(2))})", s)
            s = re.sub(r"single_long_stint\(([-\d\.]+)h\)", lambda m: f"single_long_stint({_fmt_hhmm(m.group(1))})", s)
            s = re.sub(r"(very_low_weekday|gt_daily_max)\(([-\d\.]+)\)", lambda m: f"{m.group(1)}({_fmt_hhmm(m.group(2))})", s)
            return s
        rev_disp["Reasons_pretty"] = rev_disp.apply(_pretty_reasons, axis=1)

        rev_disp = rev_disp.sort_values(
            ["PersonScore","PersonRows","Person","Date"],
            ascending=[False, False, True, True]
        )

    daily_disp = daily[["Date","Employee","RawHours","RoundedHours"]].copy()
    daily_disp["RawHours_str"]     = daily_disp["RawHours"].apply(lambda x: _fmt_hhmm(x) if pd.notna(x) else "")
    daily_disp["RoundedHours_str"] = daily_disp["RoundedHours"].apply(lambda x: _fmt_hhmm(x) if pd.notna(x) else "")

    preview_longest, preview_ot = _build_preview_compact_lists(
        rows_in_block=rows_in_block,
        hours_lookup=hours_lookup,
        tar_to_wk=tar_to_wk,
        round_to_hours=round_to_hours,
        reg_cap=reg_cap,
        long_stint_flag=long_stint_flag,
        suggest_lunch_deduct=suggest_lunch_deduct,
        min_stint_for_list=4.0,
        ot_threshold=8.0,
    )

    # 5) Fill WeeklyTime Reg/OT
    filled = 0
    for r_idx, wname in weekly_rows:
        tn, score = wk_to_tar.get(wname, (None, 0))
        if not tn:
            continue
        for d, mp in day_map.items():
            rounded = hours_lookup.get((d, tn), 0.0)
            reg = round(min(rounded, reg_cap), 2)
            ot  = round(max(rounded - reg_cap, 0.0), 2)
            if ws is not None:
                ws.cell(row=r_idx+1, column=mp["reg_col"]+1).value = reg
                ws.cell(row=r_idx+1, column=mp["ot_col"]+1).value  = ot
            filled += 1

    # 6) Longest-stint views
    longest_by_day, longest_leader = _build_longest_stint_views(
        rows_in_block=rows_in_block,
        hours_lookup=hours_lookup,
        tar_to_wk=tar_to_wk,
        long_stint_flag=long_stint_flag,
        round_to_hours=round_to_hours,
        suggest_lunch_deduct=suggest_lunch_deduct,
    )

    # 7) Save to bytes
    output_bytes = None
    if save and wb is not None:
        def _write_df(sheet_name: str, df: pd.DataFrame):
            if sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])
            ws_new = wb.create_sheet(sheet_name)
            ws_new.append([str(c) for c in df.columns])
            for _, r in df.iterrows():
                ws_new.append([
                    r.get(c, "") if isinstance(r, dict)
                    else (r[c] if c in df.columns else "")
                    for c in df.columns
                ])

        _write_df("Review_Queue", review_df)
        _write_df("Name_Matching", name_matching)
        _write_df("Daily_Hours_Long", daily[["Date","Employee","RawHours","RoundedHours"]])
        if not rev_disp.empty:
            _write_df("Review_Queue_Display", rev_disp[[
                "Date","Person","Weekly_Name","TAR_Name",
                "RawHours_str","RoundedHours_str","SuggestedHours_str","LongestStint_str",
                "Reasons_pretty","ViolationWeight","PersonScore"
            ]])
        if not violations_by_person.empty:
            _write_df("Violations_By_Person", violations_by_person)
        if longest_leader is not None and not longest_leader.empty:
            _write_df("Longest_Stints_Leaderboard", longest_leader)
        if longest_by_day is not None and not longest_by_day.empty:
            _write_df("Longest_Stints_By_Day", longest_by_day)
        if preview_longest is not None and not preview_longest.empty:
            _write_df("Preview_Longest_Shifts", preview_longest)
        if preview_ot is not None and not preview_ot.empty:
            _write_df("Preview_OT_Days", preview_ot)

        output_buffer = BytesIO()
        wb.save(output_buffer)
        output_bytes = output_buffer.getvalue()

    # 8) Build secretary summary message
    msg = build_secretary_message(
        filled_cells=filled,
        review_df=review_df,
        day_map=day_map,
        name_matching=name_matching,
        round_to_hours=round_to_hours,
        reg_cap=reg_cap
    )

    return TimeToWeeklyResult(
        output_bytes=output_bytes,
        filled_cells=filled,
        review_df=review_df,
        name_matching=name_matching,
        daily_long=daily[["Date","Employee","RawHours","RoundedHours"]],
        secretary_message=msg,
        review_df_display=rev_disp,
        daily_long_display=daily_disp,
        violations_by_person=violations_by_person,
        longest_stints_by_day=longest_by_day,
        longest_stints_leaderboard=longest_leader,
        preview_longest_shifts=preview_longest,
        preview_ot_days=preview_ot,
    )


# =======================
# Secretary message
# =======================
def build_secretary_message(
    *,
    filled_cells: int,
    review_df: pd.DataFrame,
    day_map: Dict[datetime.date, dict],
    name_matching: pd.DataFrame,
    round_to_hours: float,
    reg_cap: float,
    exclude_reasons=("missing_day_for_regular", "rounded"),
    max_examples=8
) -> str:
    dates_sorted = sorted(day_map.keys())
    week_span = f"{dates_sorted[0].strftime('%m/%d')} – {dates_sorted[-1].strftime('%m/%d')}" if dates_sorted else "N/A"

    flagged_effective = 0
    for _, r in review_df.iterrows():
        keys = _reason_keys(r.get("Reasons", ""))
        keys_no_rounding = {k for k in keys if k != "rounded"}
        suggested_present = str(r.get("SuggestedHours", "")).strip() not in ("", "None")
        rounded_val = r.get("RoundedHours", None)
        rounding_only = (not keys_no_rounding) and not suggested_present
        rounded_is_8 = (isinstance(rounded_val, (int, float)) and abs(float(rounded_val) - 8.0) < 1e-9)
        if not rounding_only or not rounded_is_8:
            if keys_no_rounding or suggested_present:
                flagged_effective += 1

    low_conf = int((name_matching["Flag"] == "REVIEW").sum()) if not name_matching.empty else 0

    reason_counts = defaultdict(int)
    if not review_df.empty:
        for s in review_df.get("Reasons", []):
            for part in _reason_keys(s):
                if part in exclude_reasons:
                    continue
                reason_counts[part] += 1
        if "low_name_match" in reason_counts:
            reason_counts["low_name_match"] = low_conf

    order_hint = ["low_name_match", "single_long_stint", "very_low_weekday", "gt_daily_max"]
    keys_sorted = sorted(
        reason_counts.keys(),
        key=lambda k: (order_hint.index(k) if k in order_hint else 999, -reason_counts[k], k)
    )
    bullets = [f"• {k.replace('_',' ')}: {reason_counts[k]}" for k in keys_sorted]

    def _display_name(rec):
        name = (rec.get("Weekly_Name") or rec.get("TAR_Name") or "").strip()
        return name

    def _looks_like_total(name):
        s = str(name).lower()
        return s.startswith("total") or "grand total" in s

    def _concerns_for_row(r):
        concerns = []
        reasons = r.get("Reasons", "")
        keys = _reason_keys(reasons)

        if "low_name_match" in keys:
            score = r.get("MatchScore", 0)
            try:
                score = int(round(float(score)))
            except Exception:
                pass
            concerns.append(f"Low Name Match: {score}%")

        segs = []
        segs_str = str(r.get("Segments", "") or "").strip()
        if segs_str:
            for tok in segs_str.split(";"):
                tok = tok.strip()
                if not tok:
                    continue
                try:
                    segs.append(float(tok))
                except:
                    pass
        if segs and "single_long_stint" in keys:
            longest = max(segs)
            concerns.append(f"Long Stint: {_fmt_hhmm(longest)}")

        sug = r.get("SuggestedHours", "")
        if isinstance(sug, (int, float)) and isinstance(r.get("RoundedHours", None), (int, float)):
            delta_m = _mins_from_hours(float(r["RoundedHours"]) - float(sug))
            if delta_m and delta_m > 0:
                concerns.append(f"Reduced {delta_m}m")

        if "very_low_weekday" in keys:
            rv = r.get("RoundedHours", None)
            if isinstance(rv, (int,float)):
                concerns.append(f"Low Weekday: {round(float(rv), 2)}h")

        if "gt_daily_max" in keys:
            rv = r.get("RoundedHours", None)
            if isinstance(rv, (int,float)):
                concerns.append(f"Over Daily Max: {_fmt_hhmm(rv)}")

        return concerns

    examples = []
    seen_names = set()
    if not review_df.empty:
        for _, r in review_df.iterrows():
            name = _display_name(r)
            if not name or _looks_like_total(name) or name in seen_names:
                continue
            concerns = _concerns_for_row(r)
            if not concerns:
                continue
            line = f"- {name} on {r['Date']}: " + ", ".join(concerns)
            examples.append(line)
            seen_names.add(name)
            if len(examples) >= max_examples:
                break

    msg_lines = [
        f"Week {week_span}: populated Reg (≤{reg_cap}h) & OT (> {reg_cap}h), rounded to nearest {round_to_hours}h.",
        f"Filled {filled_cells} day-cells.",
        f"Flagged {flagged_effective} item(s) for review." + (f"  Low-confidence name matches: {low_conf}." if low_conf else ""),
    ]
    if bullets:
        msg_lines.append("Heads-up:")
        msg_lines.extend(bullets)
    if examples:
        msg_lines.append("\nExamples:")
        msg_lines.extend(examples)
    msg_lines.append("\nOpen the 'Review_Queue' sheet to fix any items. Thanks!")
    return "\n".join(msg_lines)
