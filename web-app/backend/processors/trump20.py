# Trump20 - Daily Processor
# Adapted from original Trump20.py to work with file bytes for web uploads
# Preserves EXACT cell reading/writing logic for client's hyper-specific Excel formats

import os, re, math, datetime
from collections import defaultdict, Counter
from dataclasses import dataclass
from typing import Dict, Tuple, Optional, List
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook
from fuzzywuzzy import fuzz
from functools import lru_cache


# =========================
# Regex helpers
# =========================
TIME_HHMM_RE   = re.compile(r'^(\d+):(\d{2})(?::\d{2})?$')
DATE_HEADER_RE = re.compile(r"Timecard Date:\s*(\d{1,2}/\d{1,2}/\d{4})")
DATE_MMDD_RE = re.compile(r'(\d{1,2})/(\d{1,2})')

# ==========================
# General-purpose utils
# ==========================
def _round_to(x: float, step: float) -> float:
    return round(float(x) / step) * step

def round_half_hour_with_8_cutoff(hours):
    """Round to nearest 30 min, except 8:00–8:25 → 8:00."""
    if hours is None:
        return None
    try:
        mins = int(round(float(hours) * 60))
    except Exception:
        return hours
    EIGHT = 8 * 60
    if EIGHT <= mins <= EIGHT + 20:
        return 8.0
    rounded_mins = ((mins + 15) // 30) * 30
    return rounded_mins / 60.0

@lru_cache(maxsize=4096)
def _norm(s: str) -> str:
    return re.sub(r"[^a-z\s\-']", " ", str(s).lower()).strip()

def _best_match(needle: str, haystack: List[str], min_score: int, fallback_score: int):
    """Fuzzy match `needle` to the best item in `haystack` (strict + fallback by last name)."""
    wn = _norm(needle)
    best, score = None, -1
    for c in haystack:
        sc = fuzz.token_set_ratio(wn, _norm(c))
        if sc > score:
            best, score = c, sc
    if score >= min_score:
        return best, score

    wn_last = wn.split()[-1] if wn.split() else ""
    fallback, fscore = None, -1
    for c in haystack:
        cn = _norm(c)
        c_last = cn.split()[-1] if cn.split() else ""
        if c_last == wn_last:
            sc = fuzz.token_set_ratio(wn, cn)
            if sc > fscore:
                fallback, fscore = c, sc
    if fallback and fscore >= fallback_score:
        return fallback, fscore
    return None, 0

def _fmt_hhmm(hours) -> str:
    """Float hours → 'H:MM' (7.5 → '7:30')."""
    if hours is None or (isinstance(hours, float) and (math.isnan(hours) or math.isinf(hours))):
        return ""
    try:
        h = float(hours)
    except Exception:
        return str(hours)
    neg = h < 0
    h = abs(h)
    H = int(math.floor(h))
    M = int(round((h - H) * 60))
    if M == 60:
        H += 1; M = 0
    return f"{'-' if neg else ''}{H}:{M:02d}"

def _is_weekday(d: datetime.date) -> bool:
    return d.weekday() < 5

def _parse_hours_cell(x) -> Optional[float]:
    """Accept numbers, 'H:MM' or 'HH:MM(:SS)', or '8.0' → float hours."""
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return None if pd.isna(x) else float(x)
    s = str(x).strip()
    if not s:
        return None
    m = TIME_HHMM_RE.match(s)
    if m:
        h = int(m.group(1)); mm = int(m.group(2))
        return h + mm / 60.0
    try:
        return float(s)
    except:
        return None


# ======================================
# Parse ONE-DAY Time Activity Report
# ======================================
def parse_one_day_time_activity(raw_times_bytes: bytes):
    """
    Accepts bytes instead of file path.
    Returns:
        report_date (datetime.date)
        daily_df: DataFrame columns [Employee, RawHours, RoundedHours]
        stints_map: {Employee -> [stint1, stint2, ...]}   (floats in hours)
    """
    xl = pd.ExcelFile(BytesIO(raw_times_bytes), engine="openpyxl")
    df = xl.parse(xl.sheet_names[0], header=None)

    report_date = None
    stints_map = defaultdict(list)
    records = []

    # 1) Find the first (and only) 'Timecard Date: mm/dd/yyyy'
    for _, row in df.iterrows():
        for cell in row.tolist():
            if isinstance(cell, str):
                m = DATE_HEADER_RE.search(cell)
                if m:
                    report_date = datetime.datetime.strptime(m.group(1), "%m/%d/%Y").date()
                    break
        if report_date:
            break
    if not report_date:
        raise RuntimeError("Could not find a 'Timecard Date:' header in the one-day report.")

    # 2) Walk rows *after* the date header and collect (Employee, Total Hours)
    started = False
    for _, row in df.iterrows():
        if not started:
            if any(isinstance(cell, str) and DATE_HEADER_RE.search(cell or "") for cell in row.tolist()):
                started = True
            continue
        if isinstance(row.iloc[0], str) and row.iloc[0].strip().lower() == "employee":
            continue

        emp = row.iloc[0] if row.shape[0] > 0 else None
        if not emp or str(emp).strip().lower() == "nan":
            continue

        hrs = _parse_hours_cell(row.iloc[5] if row.shape[0] > 5 else None)
        if hrs is None or pd.isna(hrs) or hrs <= 0:
            continue

        emp_str = str(emp).strip()
        low = emp_str.lower()
        if low.startswith("total") or "grand total" in low:
            continue

        records.append({"Employee": emp_str, "RawHours": float(hrs)})
        stints_map[emp_str].append(float(hrs))

    if not records:
        raise RuntimeError("Parsed zero rows. Check that column A=Employee and column F=Total Hours.")

    daily_df = pd.DataFrame(records).groupby("Employee", as_index=False)["RawHours"].sum()
    return report_date, daily_df, stints_map


# =======================================
# Read WeeklyTime & find day/name cols
# =======================================
def read_weekly_structure(weekly_template_bytes: bytes):
    """
    Accepts bytes instead of file path.
    Returns:
        day_map: {date -> {"reg_col": j, "ot_col": j+1, "header": str}}
        weekly_rows: [(row_idx0, display_name)]
        name_col: int
        start_year: int
        wb, ws, wk_df
    """
    xl = pd.ExcelFile(BytesIO(weekly_template_bytes), engine="openpyxl")
    sheet = xl.sheet_names[0]
    wk_df = xl.parse(sheet, header=None)

    # --- Find "Week Of" row and infer year (MM.DD.YY - MM.DD.YY) ---
    week_of_row = None
    for i, v in enumerate(wk_df.iloc[:, 0]):
        if isinstance(v, str) and "week of" in v.lower():
            week_of_row = i
            break
    if week_of_row is None:
        raise RuntimeError("Couldn't find a 'Week Of :' row in WeeklyTime.")

    week_str = str(wk_df.iat[week_of_row, 0])
    m = re.search(r'(\d{2})\.(\d{2})\.(\d{2})\s*-\s*(\d{2})\.(\d{2})\.(\d{2})', week_str)
    start_year = 2000 + int(m.group(3)) if m else datetime.date.today().year

    day_hdr_row = week_of_row + 1
    sub_hdr_row = day_hdr_row + 1

    # --- Build map of date → (Reg col, OT col) by scanning headers ---
    day_map: Dict[datetime.date, dict] = {}
    for j in range(wk_df.shape[1]):
        val = wk_df.iat[day_hdr_row, j]

        mm = dd = None
        header_text = None

        if isinstance(val, str):
            m = DATE_MMDD_RE.search(val)
            if m:
                mm, dd = map(int, m.groups())
                header_text = val
        elif isinstance(val, (datetime.date, datetime.datetime, pd.Timestamp)):
            dt = pd.to_datetime(val)
            mm, dd = int(dt.month), int(dt.day)
            header_text = dt.strftime("%-m/%-d") if hasattr(dt, "strftime") else f"{mm}/{dd}"

        if mm is None or dd is None:
            continue

        reg_ok = (
            isinstance(wk_df.iat[sub_hdr_row, j], str)
            and "reg" in str(wk_df.iat[sub_hdr_row, j]).lower()
        )
        ot_ok = (
            j + 1 < wk_df.shape[1]
            and isinstance(wk_df.iat[sub_hdr_row, j + 1], str)
            and "ot" in str(wk_df.iat[sub_hdr_row, j + 1]).lower()
        )
        if not (reg_ok and ot_ok):
            continue

        day_date = datetime.date(start_year, mm, dd)
        day_map[day_date] = {"reg_col": j, "ot_col": j + 1, "header": (header_text or f"{mm}/{dd}")}

    if not day_map:
        raise RuntimeError("Couldn't map any day columns in WeeklyTime (check headers: MM/DD + Reg/OT).")

    # --- Dynamically find the 'Employee Name' column & start row ---
    name_col = None
    start_row = None
    for j in range(wk_df.shape[1]):
        col = wk_df.iloc[:, j]
        for i, v in enumerate(col):
            if isinstance(v, str) and v.strip().lower().startswith("employee name"):
                name_col = j
                start_row = i + 1
                break
        if name_col is not None:
            break
    if name_col is None or start_row is None:
        raise RuntimeError("Couldn't find 'Employee Name' header in WeeklyTime.")

    # --- Collect the employee display rows from that column ---
    weekly_rows: List[Tuple[int, str]] = []
    for r in range(start_row, wk_df.shape[0]):
        name_cell = wk_df.iat[r, name_col]
        if name_cell is None or (isinstance(name_cell, float) and pd.isna(name_cell)):
            continue
        s = str(name_cell).strip()
        if s and s.lower() != "nan":
            weekly_rows.append((r, s))

    wb = load_workbook(BytesIO(weekly_template_bytes))
    ws = wb[wb.sheetnames[0]]
    return day_map, weekly_rows, name_col, start_year, wb, ws, wk_df


# ============================================
# Result dataclass
# ============================================
@dataclass
class Trump20Result:
    output_bytes: Optional[bytes]
    report_date: datetime.date
    filled_cells: int
    review_df: pd.DataFrame
    name_matching: pd.DataFrame
    daily_long: pd.DataFrame
    secretary_message: str


def run_trump20_daily(
    raw_times_bytes: bytes,
    weekly_template_bytes: bytes,
    *,
    round_to_hours: float       = 0.5,
    reg_cap: float              = 8.0,
    daily_max_hours: float      = 16.0,
    long_stint_flag: float      = 10.0,
    match_min_score: int        = 92,
    fallback_score: int         = 85,
    flag_low_weekday: float     = 2.0,
    suggest_lunch_deduct: float = 0.5,
) -> Trump20Result:
    """
    Process a single day's Time Activity Report and fill the corresponding day in WeeklyTime.
    Accepts bytes instead of file paths for web compatibility.
    """
    # ---- 1) Parse the one-day report ----
    report_date, daily_df, stints_map = parse_one_day_time_activity(raw_times_bytes)
    daily_df = daily_df.copy()
    daily_df["RoundedHours"] = daily_df["RawHours"].apply(round_half_hour_with_8_cutoff)

    # ---- 2) Read WeeklyTime structure ----
    day_map, weekly_rows, name_col, start_year, wb, ws, wk_df = read_weekly_structure(weekly_template_bytes)

    target_date = report_date
    if report_date not in day_map:
        mmdd_map = {(d.month, d.day): d for d in day_map.keys()}
        key = (report_date.month, report_date.day)
        if key in mmdd_map:
            target_date = mmdd_map[key]
        else:
            have = ", ".join(sorted({f"{d.month}/{d.day}" for d in day_map.keys()}))
            raise RuntimeError(
                f"The date {report_date.strftime('%m/%d/%Y')} does not exist in this WeeklyTime file. "
                f"Found day headers for: {have}."
            )

    # ---- 3) Name matching (TAR names → WeeklyTime names) ----
    weekly_names = [n for _, n in weekly_rows]
    tar_names = sorted(daily_df["Employee"].astype(str).unique())

    match_rows = []
    tar_to_wk: Dict[str, Tuple[str, int]] = {}
    wk_to_tar: Dict[str, Tuple[str, int]] = {}

    for tn in tar_names:
        wmatch, score = _best_match(tn, weekly_names, match_min_score, fallback_score)
        match_rows.append({
            "TAR Name": tn,
            "Weekly Match": wmatch or "",
            "Score": score,
            "Flag": "" if score >= match_min_score else "REVIEW"
        })
        if wmatch:
            tar_to_wk[tn] = (wmatch, score)

    for tn, (wmatch, score) in tar_to_wk.items():
        if wmatch not in wk_to_tar or score > wk_to_tar[wmatch][1]:
            wk_to_tar[wmatch] = (tn, score)

    name_matching = pd.DataFrame(match_rows).sort_values(
        ["Flag","Score","Weekly Match","TAR Name"], ascending=[True, False, True, True]
    )

    # ---- 4) Build review queue for THIS day only ----
    review = []
    hours_lookup = {tn: float(h) for tn, h in zip(daily_df["Employee"], daily_df["RoundedHours"])}

    def add_review_row(tn, wmatch, score, raw_val, rounded_val, stints, reasons, suggested):
        review.append({
            "Date": report_date.strftime("%m/%d/%Y"),
            "TAR_Name": tn,
            "Weekly_Name": wmatch or "",
            "MatchScore": score,
            "Segments": ";".join(f"{s:.2f}" for s in stints) if stints else "",
            "RawHours": round(raw_val, 2) if raw_val is not None else "",
            "RoundedHours": round(rounded_val, 2) if rounded_val is not None else "",
            "SuggestedHours": suggested if suggested is not None else "",
            "Reasons": ", ".join(reasons) if reasons else ""
        })

    for _, row in daily_df.iterrows():
        tn = row["Employee"]
        rounded = float(row["RoundedHours"])
        raw = float(row["RawHours"])
        wmatch, score = tar_to_wk.get(tn, (None, 0))

        reasons, suggested = [], None
        if score < match_min_score:
            reasons.append(f"low_name_match({score})")
        if rounded > daily_max_hours:
            reasons.append(f"gt_daily_max({rounded})")
        if rounded > 0 and rounded <= flag_low_weekday and _is_weekday(report_date):
            reasons.append(f"very_low_weekday({rounded})")
        if abs(raw - rounded) >= 0.01:
            reasons.append(f"rounded({raw:.2f}->{rounded:.2f})")

        stints = stints_map.get(tn, [])
        if len(stints) == 1 and stints[0] >= long_stint_flag:
            reasons.append(f"single_long_stint({stints[0]:.2f}h)")
            suggested = round_half_hour_with_8_cutoff(max(rounded - suggest_lunch_deduct, 0.0))

        if reasons or suggested is not None:
            add_review_row(tn, wmatch, score, raw, rounded, stints, reasons, suggested)

    review_df = pd.DataFrame(review).sort_values(["Weekly_Name","TAR_Name"]).reset_index(drop=True)

    # ---- 5) Fill ONLY this day's Reg/OT in WeeklyTime ----
    filled = 0
    reg_col = day_map[target_date]["reg_col"]
    ot_col  = day_map[target_date]["ot_col"]

    name_to_row = {name: r for r, name in weekly_rows}

    for wname, (tn, score) in wk_to_tar.items():
        rounded = float(hours_lookup.get(tn, 0.0))
        reg = round(min(rounded, reg_cap), 2)
        ot  = round(max(rounded - reg_cap, 0.0), 2)

        r_idx0 = name_to_row.get(wname, None)
        if r_idx0 is None:
            continue

        reg_cell = ws.cell(row=r_idx0 + 1, column=reg_col + 1)
        ot_cell  = ws.cell(row=r_idx0 + 1, column=ot_col  + 1)

        reg_cell.value = 0 if reg <= 0 else reg
        ot_cell.value = None if ot <= 0 else ot
        filled += 2

    # ---- 6) Save to bytes ----
    output_buffer = BytesIO()
    wb.save(output_buffer)
    output_bytes = output_buffer.getvalue()

    # ---- 7) Build secretary message ----
    daily_totals = {str(r["Employee"]): float(r["RoundedHours"]) for _, r in daily_df.iterrows()}

    unmatched_tar = sorted([tn for tn in tar_names if tn not in tar_to_wk])

    long_shift_rows = []
    for tn, stints in stints_map.items():
        if not stints:
            continue
        longest = max([s for s in stints if isinstance(s, (int, float)) and not math.isnan(s)] or [0.0])
        if longest > 10.0:
            person = (tar_to_wk.get(tn, (None, 0))[0] or tn)
            total = daily_totals.get(tn, 0.0)
            long_shift_rows.append((person, longest, total))
    long_shift_rows.sort(key=lambda x: (-x[1], x[0]))

    short_day_rows = []
    for tn, total in daily_totals.items():
        if total > 0.5 and total < 4.0:
            person = (tar_to_wk.get(tn, (None, 0))[0] or tn)
            short_day_rows.append((person, total))
    short_day_rows.sort(key=lambda x: x[0].lower())

    lines = [
        f"Updated {report_date.strftime('%a %m/%d/%Y')}: wrote Reg/OT for {len(wk_to_tar)} matched employee(s).",
        f"Rounded to nearest {round_to_hours}h; Reg cap {reg_cap}h.",
    ]

    if unmatched_tar:
        lines += ["", f"Not in WeeklyTime ({len(unmatched_tar)}):"]
        lines += [f"• {n}" for n in unmatched_tar]

    if long_shift_rows:
        lines += ["", f"Long shifts >10:00 ({len(long_shift_rows)}):"]
        for person, longest, total in long_shift_rows:
            lines.append(f"• {person} — longest {_fmt_hhmm(longest)} (total {_fmt_hhmm(total)})")

    if short_day_rows:
        lines += ["", f"Very short day (>0:30 and <4:00) ({len(short_day_rows)}):"]
        for person, total in short_day_rows:
            lines.append(f"• {person} — total {_fmt_hhmm(total)}")

    secretary_message = "\n".join(lines)

    return Trump20Result(
        output_bytes=output_bytes,
        report_date=report_date,
        filled_cells=filled,
        review_df=review_df,
        name_matching=name_matching,
        daily_long=daily_df[["Employee","RawHours","RoundedHours"]],
        secretary_message=secretary_message,
    )
